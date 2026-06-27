"""
SpectraSoft — Job 5: INT.1 (Raw Intensity) - Simulation Mode

This page simulates reading data from an Excel file.
User enters an ST number, and the software reads the corresponding column
from an Excel file (simulation_data.xlsx) to populate the results table.

Element names are fetched from the first column of the Excel file.
Columns: Element, AVE, N=1, N=2, ...
No R, S.D., C.V. columns.

Export as Excel (.xlsx) using openpyxl.
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QFrame, QMessageBox, QLineEdit,
    QProgressBar, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QFileDialog
)
from PyQt6.QtCore import Qt, QTimer, QDate
from PyQt6.QtGui import QColor, QTextDocument, QPageLayout
from PyQt6.QtPrintSupport import QPrinter, QPrintPreviewDialog

import os
import math

# Use openpyxl for both reading and writing
try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Default simulation Excel file
SIMULATION_EXCEL = "simulation_data.xlsx"


class Job5RunPage(QWidget):
    """Job 5: Raw Intensity simulation page reading from Excel."""

    def __init__(self, main_window, group_id: int, group_name: str, job_type: str):
        super().__init__()
        self.main_window = main_window
        self.group_id = group_id
        self.group_name = group_name
        self.job_type = job_type   # Should be '5'

        self.results = []          # list of dict: each burn result {element: intensity}
        self.element_names = []    # Will be populated from Excel file
        self.is_running = False

        # Counters for status footer
        self.an_count = 0
        self.tan_count = 0

        self.setAutoFillBackground(True)
        p = self.palette()
        p.setColor(self.backgroundRole(), Qt.GlobalColor.lightGray)
        self.setPalette(p)

        self._build_ui()
        self._setup_job_ui()
        self._load_elements_from_excel()  # Load elements from Excel on start
        self._update_table()

    # =========================================================================
    # UI Construction
    # =========================================================================

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Title Bar ──────────────────────────────────────────────────────
        title_bar = QWidget()
        title_bar.setFixedHeight(24)
        title_bar.setStyleSheet("background:#5c9bd5;")
        title_layout = QHBoxLayout(title_bar)
        title_layout.setContentsMargins(12, 0, 12, 0)

        self.title_label = QLabel(f"Job 5: Raw Intensity (Simulation) - {self.group_name}")
        self.title_label.setStyleSheet("color:white;font:bold 10pt Arial;")
        title_layout.addWidget(self.title_label)
        title_layout.addStretch()

        self.hv_btn = QPushButton("HV: OFF")
        self.hv_btn.setStyleSheet(
            "QPushButton{"
            "background:#dc3545;"
            "color:white;"
            "border:2px outset #888888;"
            "font:9pt Arial;"
            "padding:2px 8px;"
            "}"
        )
        self.hv_btn.setFixedWidth(80)
        self.hv_btn.clicked.connect(self._toggle_hv)
        title_layout.addWidget(self.hv_btn)

        root.addWidget(title_bar)

        # ── Outer Frame ──────────────────────────────────────────────────
        outer = QFrame()
        outer.setFrameShape(QFrame.Shape.Box)
        outer.setFrameShadow(QFrame.Shadow.Sunken)
        outer.setLineWidth(2)
        outer.setStyleSheet("background:white;")
        root.addWidget(outer, stretch=1)

        ol = QVBoxLayout(outer)
        ol.setContentsMargins(10, 10, 10, 10)
        ol.setSpacing(6)

        # ── Job Parameters Area (ST Number) ────────────────────────────
        self.params_area = QWidget()
        self.params_area.setFixedHeight(40)
        ol.addWidget(self.params_area)

        # ── Progress Bar ─────────────────────────────────────────────────
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet(
            "QProgressBar{"
            "background:#f0f0f0;"
            "border:1px solid #888888;"
            "height:20px;"
            "text-align:center;"
            "}"
            "QProgressBar::chunk{"
            "background:#0078d7;"
            "}"
        )
        self.progress_bar.setValue(0)
        ol.addWidget(self.progress_bar)

        # ── Status Label ─────────────────────────────────────────────────
        self.status_label = QLabel("Ready. Enter ST number and press F1: Start.")
        self.status_label.setStyleSheet(
            "QLabel{"
            "background:#d4d0c8;"
            "color:#333333;"
            "font:9pt Arial;"
            "border:1px solid #888888;"
            "padding:4px 6px;"
            "}"
        )
        ol.addWidget(self.status_label)

        # ── ST Counter ──────────────────────────────────────────────────
        self.st_counter = QLabel("ST No.: —")
        self.st_counter.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.st_counter.setStyleSheet("font:bold 9pt Arial; color:#555555;")
        ol.addWidget(self.st_counter)

        # ── Results Table ────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setStyleSheet(
            "QTableWidget{"
            "background:white;"
            "color:black;"
            "border:1px solid #888888;"
            "gridline-color:#888888;"
            "font:9pt Arial;"
            "}"
            "QTableWidget::item{"
            "border:1px solid #888888;"
            "padding:0px 4px;"
            "color:black;"
            "background:white;"
            "}"
            "QHeaderView::section{"
            "background:#0078d7;"
            "color:white;"
            "font:bold 9pt Arial;"
            "border:1px solid #888888;"
            "padding:2px 4px;"
            "}"
            "QTableWidget::item:selected{"
            "background:#cce5ff;"
            "color:black;"
            "}"
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setDefaultSectionSize(27)
        ol.addWidget(self.table, stretch=1)

        # ── Status Footer ──────────────────────────────────────────────────
        footer = QWidget()
        footer.setAutoFillBackground(True)
        fbp = footer.palette()
        fbp.setColor(footer.backgroundRole(), Qt.GlobalColor.lightGray)
        footer.setPalette(fbp)
        footer.setStyleSheet("background:#d4d0c8; border-top:1px solid #888888;")
        footer.setFixedHeight(28)

        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(10, 4, 10, 4)
        footer_layout.setSpacing(20)

        self.an_label = QLabel("AN: 0")
        self.an_label.setStyleSheet("font:9pt Arial; color:black;")
        footer_layout.addWidget(self.an_label)

        self.tan_label = QLabel("TAN: 0")
        self.tan_label.setStyleSheet("font:9pt Arial; color:black;")
        footer_layout.addWidget(self.tan_label)

        self.hv_status_label = QLabel("HV: OFF")
        self.hv_status_label.setStyleSheet("font:9pt Arial; color:red;")
        footer_layout.addWidget(self.hv_status_label)

        footer_layout.addStretch()
        ol.addWidget(footer)

        # ── Bottom Navigation ────────────────────────────────────────────
        btn_bar = QWidget()
        btn_bar.setAutoFillBackground(True)
        bbp = btn_bar.palette()
        bbp.setColor(btn_bar.backgroundRole(), Qt.GlobalColor.lightGray)
        btn_bar.setPalette(bbp)
        btn_bar.setFixedHeight(40)

        bbl = QHBoxLayout(btn_bar)
        bbl.setContentsMargins(0, 4, 0, 4)
        bbl.setSpacing(4)

        btn_style = (
            "QPushButton{"
            "background:#d4d0c8;"
            "color:black;"
            "border:2px outset #aaaaaa;"
            "font:9pt Arial;"
            "padding:4px 12px;"
            "min-width:60px;"
            "}"
            "QPushButton:pressed{"
            "border:2px inset #888888;"
            "}"
        )

        self.btn_start = QPushButton("F1: Start")
        self.btn_start.setStyleSheet(btn_style)
        self.btn_start.clicked.connect(self._on_start)

        self.btn_stop = QPushButton("F2: Stop")
        self.btn_stop.setStyleSheet(btn_style)
        self.btn_stop.clicked.connect(self._on_stop)
        self.btn_stop.setEnabled(False)

        self.btn_end = QPushButton("F5: End Analysis")
        self.btn_end.setStyleSheet(btn_style)
        self.btn_end.clicked.connect(self._on_end_analysis)

        self.btn_print = QPushButton("F4: Print")
        self.btn_print.setStyleSheet(btn_style)
        self.btn_print.clicked.connect(self._on_print)

        self.btn_export = QPushButton("Export Excel")
        self.btn_export.setStyleSheet(btn_style)
        self.btn_export.clicked.connect(self._on_export)

        bbl.addWidget(self.btn_start)
        bbl.addWidget(self.btn_stop)
        bbl.addWidget(self.btn_end)
        bbl.addWidget(self.btn_print)
        bbl.addWidget(self.btn_export)
        bbl.addStretch()

        canc = QPushButton("9:Cancel")
        canc.setStyleSheet(btn_style)
        canc.clicked.connect(self._on_cancel)
        bbl.addWidget(canc)

        ol.addWidget(btn_bar)

    # =========================================================================
    # Job-Specific UI Setup
    # =========================================================================

    def _setup_job_ui(self):
        """Set up job-specific parameters UI (ST Number input)."""
        if self.params_area.layout():
            while self.params_area.layout().count():
                item = self.params_area.layout().takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            old_layout = self.params_area.layout()
            if old_layout:
                old_layout.deleteLater()

        layout = QVBoxLayout(self.params_area)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        row = QHBoxLayout()
        row.setSpacing(8)

        lbl = QLabel("ST Number:")
        lbl.setStyleSheet("color:black;font:9pt Arial;")
        row.addWidget(lbl)

        self.st_number_input = QLineEdit()
        self.st_number_input.setPlaceholderText("e.g., abc123")
        self.st_number_input.setStyleSheet(
            "QLineEdit{background:white;color:black;border:1px solid #888888;"
            "font:9pt Arial;padding:2px 4px;}"
        )
        self.st_number_input.setFixedWidth(200)
        row.addWidget(self.st_number_input)

        row.addStretch()
        layout.addLayout(row)

    # =========================================================================
    # Data Loading – Elements from Excel
    # =========================================================================

    def _load_elements_from_excel(self):
        """
        Load element names from the first column of the Excel file.
        If the file is not found or empty, fall back to an empty list.
        """
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is not installed.\n"
                "Please install it using: pip install openpyxl")
            self.element_names = []
            return

        if not os.path.exists(SIMULATION_EXCEL):
            # File missing: we can show a warning but allow the user to create it.
            self.status_label.setText(f"Excel file not found: {SIMULATION_EXCEL}")
            self.element_names = []
            return

        try:
            wb = load_workbook(SIMULATION_EXCEL, data_only=True)
            ws = wb.active

            # Read header row to validate, but we only need the first column
            header = []
            for cell in ws[1]:
                header.append(cell.value if cell.value is not None else "")

            if not header:
                self.element_names = []
                return

            # First column should be "Element"
            if header[0].strip().upper() != "ELEMENT":
                QMessageBox.critical(self, "Error",
                    "First column must be 'Element'.\n"
                    f"Found: {header[0]}")
                self.element_names = []
                return

            # Read all rows starting from row 2
            elements = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) == 0 or row[0] is None:
                    continue
                element = str(row[0]).strip()
                if element:
                    elements.append(element)

            self.element_names = elements
            self.status_label.setText(f"Loaded {len(elements)} elements from Excel.")

        except Exception as e:
            QMessageBox.critical(self, "Error Reading Excel", str(e))
            self.element_names = []

    def _load_excel_data(self, st_number: str):
        """
        Load the Excel file and extract:
        - A list of element names (first column)
        - A dict mapping element name -> value for the given ST column.
        Returns (element_names, values_dict) or (None, None) on error.
        """
        if not OPENPYXL_AVAILABLE:
            return None, None

        if not os.path.exists(SIMULATION_EXCEL):
            return None, None

        try:
            wb = load_workbook(SIMULATION_EXCEL, data_only=True)
            ws = wb.active

            # Read header row
            header = []
            for cell in ws[1]:
                header.append(cell.value if cell.value is not None else "")

            if not header:
                return None, None

            # First column must be "Element"
            if header[0].strip().upper() != "ELEMENT":
                return None, None

            # Find ST column index
            col_index = -1
            for i, col in enumerate(header):
                if col is not None and col.strip().upper() == st_number.strip().upper():
                    col_index = i
                    break

            if col_index == -1:
                return None, None

            # Read data rows
            elements = []
            values = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) == 0 or row[0] is None:
                    continue
                element = str(row[0]).strip()
                if not element:
                    continue
                if col_index >= len(row):
                    continue
                val = row[col_index]
                if val is None:
                    val = 0.0
                if isinstance(val, str):
                    val = val.replace('%', '').strip()
                try:
                    value = float(val)
                except (ValueError, TypeError):
                    value = 0.0
                elements.append(element)
                values[element] = value

            return elements, values

        except Exception as e:
            QMessageBox.critical(self, "Error Reading Excel", str(e))
            return None, None

    # =========================================================================
    # Table Management
    # =========================================================================

    def _update_table(self):
        """Update the results table with Element, AVE, N=1, N=2, ... only."""
        if not self.element_names:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        num_burns = len(self.results)
        # Columns: Element, AVE, N=1, N=2, ...
        num_cols = 2 + num_burns

        self.table.setRowCount(len(self.element_names))
        self.table.setColumnCount(num_cols)

        headers = ["Element", "AVE"]
        for i in range(num_burns):
            headers.append(f"N={i+1}")

        self.table.setHorizontalHeaderLabels(headers)

        # Set column widths
        self.table.setColumnWidth(0, 70)
        self.table.setColumnWidth(1, 70)
        for i in range(num_burns):
            self.table.setColumnWidth(2 + i, 70)

        for row, elem in enumerate(self.element_names):
            # Element name
            elem_item = QTableWidgetItem(elem)
            elem_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 0, elem_item)

            if num_burns == 0:
                # FIX: Explicitly clear the AVE column when there are no burns
                empty_item = QTableWidgetItem("")
                empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                self.table.setItem(row, 1, empty_item)
            else:
                values = []
                for burn in self.results:
                    val = burn.get(elem, 0.0)
                    values.append(val)

                if values:
                    avg = sum(values) / len(values)
                    avg_item = QTableWidgetItem(f"{avg:.3f}")
                    avg_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                    self.table.setItem(row, 1, avg_item)

                    for i, val in enumerate(values):
                        val_item = QTableWidgetItem(f"{val:.3f}")
                        val_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                        self.table.setItem(row, 2 + i, val_item)

        self.table.resizeRowsToContents()

    # =========================================================================
    # HV Controls (visual only)
    # =========================================================================

    def _toggle_hv(self):
        self.main_window.toggle_hv()
        self._update_hv_button()
        self._update_footer_hv()

    def _update_hv_button(self):
        if self.main_window.get_hv_status():
            self.hv_btn.setText("HV: ON")
            self.hv_btn.setStyleSheet(
                "QPushButton{"
                "background:#28a745;"
                "color:white;"
                "border:2px outset #888888;"
                "font:9pt Arial;"
                "padding:2px 8px;"
                "}"
            )
        else:
            self.hv_btn.setText("HV: OFF")
            self.hv_btn.setStyleSheet(
                "QPushButton{"
                "background:#dc3545;"
                "color:white;"
                "border:2px outset #888888;"
                "font:9pt Arial;"
                "padding:2px 8px;"
                "}"
            )

    def _update_footer_hv(self):
        if self.main_window.get_hv_status():
            self.hv_status_label.setText("HV: ON")
            self.hv_status_label.setStyleSheet("font:9pt Arial; color:green;")
        else:
            self.hv_status_label.setText("HV: OFF")
            self.hv_status_label.setStyleSheet("font:9pt Arial; color:red;")

    def _update_footer_counts(self):
        self.an_label.setText(f"AN: {self.an_count}")
        self.tan_label.setText(f"TAN: {self.tan_count}")

    # =========================================================================
    # Actions
    # =========================================================================

    def _on_start(self):
        """Simulate a burn: read data from Excel for the entered ST number."""
        st_number = self.st_number_input.text().strip()
        if not st_number:
            QMessageBox.warning(self, "Missing ST", "Please enter an ST number.")
            return

        elements, data = self._load_excel_data(st_number)
        if elements is None or data is None:
            QMessageBox.critical(self, "Error",
                f"Failed to load data for ST '{st_number}'.\n"
                f"Check that the Excel file exists and the ST column is present.")
            return

        # Update element names from Excel
        self.element_names = elements
        self._update_table()

        # Build burn result: only elements that exist in Excel
        burn_result = {}
        for elem in self.element_names:
            burn_result[elem] = data.get(elem, 0.0)

        # Store result
        self.results.append(burn_result)
        self.an_count += 1
        self.tan_count += 1
        self._update_footer_counts()
        self.st_counter.setText(f"ST No.: {len(self.results)}")

        self._update_table()
        self.status_label.setText(f"Burn {len(self.results)} complete (simulation).")
        self.progress_bar.setValue(100)

        QTimer.singleShot(2000, lambda: self.progress_bar.setValue(0))

    def _on_stop(self):
        self.status_label.setText("Stopped (simulation).")

    def _on_end_analysis(self):
        """Clear all burns, reset ST number input, reset counters."""
        if not self.results:
            QMessageBox.information(self, "No Data", "No burns to clear.")
            return

        reply = QMessageBox.question(
            self,
            "End Analysis",
            "Clear all burns for this sample?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            # Clear results and reset table
            self.results = []
            self.st_counter.setText("ST No.: —")
            self._update_table()  # Shows only Element and AVE (empty) columns

            # Reset counters
            self.an_count = 0
            self.tan_count = 0
            self._update_footer_counts()

            # Clear the ST number input field
            self.st_number_input.clear()

            # Reset progress and status
            self.progress_bar.setValue(0)
            self.status_label.setText("Analysis ended. Ready for new sample.")

            # Reset the HV status just to keep consistent (visual)
            self._update_footer_hv()    
            
    def _on_print(self):
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        preview = QPrintPreviewDialog(printer, self)
        preview.paintRequested.connect(self._render_print_page)
        preview.exec()

    def _render_print_page(self, printer):
        doc = QTextDocument()
        html = self._generate_html_report()
        doc.setHtml(html)
        doc.print(printer)

    def _generate_html_report(self) -> str:
        title = f"SpectraSoft Job 5 Report - {self.group_name}"
        timestamp = QDate.currentDate().toString("dd-MM-yyyy")
        st_number = self.st_number_input.text().strip() if hasattr(self, 'st_number_input') else "Unknown"

        rows = []
        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            for row in range(self.table.rowCount()):
                row_html = "<tr>"
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    val = item.text() if item else ""
                    row_html += f"<td>{val}</td>"
                row_html += "</tr>"
                rows.append(row_html)

        table_html = f"""
        <table border="1" cellpadding="5" cellspacing="0">
            <thead>
                <tr>
                    {''.join(f'<th>{self.table.horizontalHeaderItem(i).text()}</th>' for i in range(self.table.columnCount()))}
                </tr>
            </thead>
            <tbody>
                {''.join(rows)}
            </tbody>
        </table>
        """

        return f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #2c3e50; }}
                .meta {{ margin-bottom: 20px; }}
                .meta td {{ padding: 2px 10px; }}
                table {{ border-collapse: collapse; width: 100%; font-size: 10pt; }}
                th {{ background-color: #34495e; color: white; padding: 6px 8px; }}
                td {{ padding: 4px 8px; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <table class="meta">
                <tr><td><b>ST Number:</b> {st_number}</td>
                    <td><b>Date:</b> {timestamp}</td></tr>
                <tr><td><b>Job:</b> 5 (INT.1 Simulation)</td>
                    <td><b>Burns:</b> {len(self.results)}</td></tr>
            </table>
            {table_html}
            <p style="margin-top:20px;font-size:9pt;color:#666;">
                Generated by SpectraSoft
            </p>
        </body>
        </html>
        """

    # =========================================================================
    # Export to Excel
    # =========================================================================

    def _on_export(self):
        if not self.results:
            QMessageBox.warning(self, "No Data", "No burns to export.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is not installed.\n"
                "Please install it using: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Analysis Results", "analysis_results.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Results"

            headers = []
            for col in range(self.table.columnCount()):
                item = self.table.horizontalHeaderItem(col)
                headers.append(item.text() if item else "")
            ws.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save(path)
            QMessageBox.information(self, "Exported",
                f"Results exported to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    # =========================================================================
    # Navigation
    # =========================================================================

    def _on_cancel(self):
        self.main_window._show_home_content()

    def wants_fullscreen(self) -> bool:
        return True